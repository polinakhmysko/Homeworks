from airflow import DAG
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.operators.dummy import DummyOperator
from airflow.operators.python import PythonOperator
from datetime import datetime
from openpyxl import load_workbook
import time


default_args = {
    'owner': "kirill",
}


def read_xlsx_and_load_to_posrgtres(**kwargs):
    wb = load_workbook(filename='/opt/airflow/dags/data/users_data.xlsx')
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    print(data_rows)

    pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')
    conn = pg_hook.get_conn()
    cursor = conn.cursor()

    insert_sql = """
        INSERT INTO raw.users (name, surname, city, age, card_number)
        VALUES (%s, %s, %s, %s, %s);
    """

    for row in data_rows:
        name, surname, city, age, card_number = row
        cursor.execute(insert_sql, (name, surname, city, age, card_number))

    conn.commit()
    cursor.close()
    conn.close()
    
    print("[Данные вставились]")


def read_xlsx_and_load_to_posrgtres_orders(**kwargs):
    wb = load_workbook(filename='/opt/airflow/dags/data/orders_data.xlsx')
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    print(data_rows)

    pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')
    conn = pg_hook.get_conn()
    cursor = conn.cursor()

    insert_sql = """
        INSERT INTO raw.orders (id, name, amount, price, total_price, card_number)
        VALUES (%s, %s, %s, %s, %s, %s);
    """

    for row in data_rows:
        id, name, amount, price, total_price, card_number = row
        cursor.execute(insert_sql, (id, name, amount, price, total_price, card_number))

    conn.commit()
    cursor.close()
    conn.close()

def read_xlsx_and_load_to_posrgtres_deliveries(**kwargs):
    wb = load_workbook(filename='/opt/airflow/dags/data/deliveries_data.xlsx')
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    print(data_rows)

    pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')
    conn = pg_hook.get_conn()
    cursor = conn.cursor()

    insert_sql = """
        INSERT INTO raw.deliveries (id_delivery, id_order, product, company, cost, name, phone_number, beggining_time, ending_time, city, warehouse, warehouse_address)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
    """

    for row in data_rows:
        id_delivery, id_order, product, company, cost, name, phone_number, beggining_time, ending_time, city, warehouse, warehouse_address = row
        cursor.execute(insert_sql, (id_delivery, id_order, product, company, cost, name, phone_number, beggining_time, ending_time, city, warehouse, warehouse_address))

    conn.commit()
    cursor.close()
    conn.close()

with DAG(
    'import_data_from_excel',
    default_args = default_args,
    description = 'Интеграция данных из xlsx',
    schedule_interval = None,
    start_date=datetime(2025, 6, 1),
) as dag:
    d = DummyOperator(task_id="dummy")

    read_and_load_task = PythonOperator(
        task_id = 'read_and_load_users',
        python_callable=read_xlsx_and_load_to_posrgtres
    )

    read_and_load_orders_task = PythonOperator(
        task_id = 'read_and_load_orders',
        python_callable=read_xlsx_and_load_to_posrgtres_orders
    )

    read_and_load_deliveries_task = PythonOperator(
        task_id = 'read_and_load_deliveries',
        python_callable=read_xlsx_and_load_to_posrgtres_deliveries
    )

    d >> [read_and_load_task, read_and_load_orders_task, read_and_load_deliveries_task]